import json
import pandas as pd
import cleartax as ct
# import json_creator as jc
import config as cg
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os
import glob
import shutil
import datetime
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle

def item_creation_einv(r):
    items = []
    # print(r)
    # for j in range(0, len(r)):
    if float(r["IGST Amount"]) != float(0.00):
        rate = float( (float(r["IGST Amount"]) / (float(r["Item Price"])) )*100)
    elif (float(r["SGST Amount"])!= float(0)) and (float(r["CGST Amount"])!= float(0)):
        rate = float( ((float(r["SGST Amount"]) + float(r["CGST Amount"])) / (float(r["Item Price"])) )*100)
    else:
        rate = 0.0

    if r["Item Type Goods Service"] == "S":
        service = "Y"
    else:
        service = "N"
    try:
        item = {
                "SlNo": f"{1}",
                "PrdDesc": "Commission", #r["Product Type"],
                "IsServc": service,#r[""],
                "HsnCd": int(r["HSN Code"]),
                "Barcde": "", #row["ItemsList_Barcde"],
                "Qty": "",#float(r[7]) if str(r[7]) and str(r[7]).strip() else "",
                "FreeQty": "",
                "Unit": "",
                "UnitPrice": float(r["Item Price"]) if pd.notna(r["Item Price"]) else "",
                "TotAmt": float(r["Item Price"]) if pd.notna(r["Item Price"]) else "",
                "Discount": 0,#r[12] if str(r[12]) and str(r[12]).strip() else int(0),
                "PreTaxVal": "",#r[13] if str(r[13]) and str(r[13]).strip() else int(0),
                "AssAmt": float(r["Item Price"]) if pd.notna(r["Item Price"]) else "",
                "GstRt": round(rate), #r[15] if str(r[15]) and str(r[15]).strip() else int(0),
                "IgstAmt": float(r["IGST Amount"]) if str(r["IGST Amount"]) and str(r["IGST Amount"]).strip() else int(0),
                "CgstAmt": float(r["CGST Amount"]) if str(r["CGST Amount"]) and str(r["CGST Amount"]).strip() else int(0),
                "SgstAmt": float(r["SGST Amount"]) if str(r["SGST Amount"]) and str(r["SGST Amount"]).strip() else int(0),
                "CesRt": 0,#r[19] if str(r[19]) and str(r[19]).strip() else int(0),
                "CesAmt": 0,#r[20] if (r[20] is not None) else int(0),
                "CesNonAdvlAmt": 0,#r[21] if str(r[21]) and str(r[21]).strip() else int(0),
                "StateCesRt": 0,#r[22] if str(r[22]) and str(r[22]).strip() else int(0),
                "StateCesAmt": 0,#r[23] if str(r[23]) and str(r[23]).strip() else int(0),
                "StateCesNonAdvlAmt": 0,#r[24] if str(r[24]) and str(r[24]).strip() else int(0),
                "OthChrg": 0,#r[25] if str(r[25]) and str(r[25]).strip() else int(0),
                "TotItemVal": float(r["Invoice Amount"]) if pd.notna(r["Invoice Amount"]) else int(0),
                "OrdLineRef": "",#r[27],
                "OrgCntry": "",#r[28],
                "PrdSlNo": "",#r[29]
            }
    except Exception as e: 
        ct.error_logger.error("Error while creating the item file:"+str(e))
    items.append(item)
    # print(items)
    return items


def create_json(t):
    try:
        # print("create_json")
        
        # print(t[i])
        customer_no = t["Sell To Customer"]
        buyer_gstin = t["GST Registration No"]
        value = t["Supplier Address"].split(' ')
        seller_loc = value[len(value) - 1]
        items = item_creation_einv(t)
        doc_date = pd.Timestamp(t["Posting Date"]).strftime('%d/%m/%Y')
        # print(doc_date)
        buyer = ct.get_buyer(t["GST Registration No"])
        # print(t["Cust Email ID"].replace(";",","))
        # receiver_email = email_count(t["Cust Email ID"].replace(";",","))
        # print(receiver_email)
        if "SPIN" in (t["Invoice No"])[:4]:
            inv_type = "INV"
        elif "SPCM" in (t["Invoice No"])[:4]:
            inv_type = "CRN"
        elif "SPDM" in (t["Invoice No"])[:4]:
            inv_type = "DBN"
        else:
            inv_type = ""
        try:
            try:
                parts = (str(buyer["pradr"]["addr"]["flno"]) if buyer["pradr"]["addr"]["flno"] != None else "") + "," + (str(buyer["pradr"]["addr"]["bnm"]) if buyer["pradr"]["addr"]["bnm"] != None else "") + "," + (str(buyer["pradr"]["addr"]["bno"]) if buyer["pradr"]["addr"]["bno"] != None else "") + "," + (str(buyer["pradr"]["addr"]["st"]) if buyer["pradr"]["addr"]["st"] != None else "")
                parts = parts.replace("None","")
                parts = parts.split(',')
                add1 = ""
                for part in parts:
                    if len(add1) + len(part) + 1 >= 100:  # +1 for the comma
                        break
                    if add1:
                        add1 += ","
                    add1 += part
                parts = (str(buyer["pradr"]["addr"]["loc"]) if buyer["pradr"]["addr"]["loc"] != None else "") + "," + (str(buyer["pradr"]["addr"]["city"])  if buyer["pradr"]["addr"]["city"] != None else "") + "," + (str(buyer["pradr"]["addr"]["dst"]) if buyer["pradr"]["addr"]["dst"] != None else "") + "," + (str(buyer["pradr"]["addr"]["stcd"]) if buyer["pradr"]["addr"]["stcd"] != None else "") 
                parts = parts.replace("None","")
                parts = parts.split(',')
                add2 = ""
                for part in parts:
                    if len(add2) + len(part) + 1 >= 100:  # +1 for the comma
                        break
                    if add2:
                        add2 += ","
                    add2 += part
                loc = buyer["pradr"]["addr"]["loc"]
                pin = buyer["pradr"]["addr"]["pncd"]
                print(add2)
                print(add1)
                print(len(add2))
                print(len(add1))
            except:
                add1 = ""
                add2 = ""
                loc = ""
                pin = ""
            invoice = {
                        "transaction": {
                            "Version": "1.1",
                            "TranDtls": {
                                "TaxSch": "GST",
                                "SupTyp": "B2B",
                                "RegRev": "N",
                                "EcmGstin": None,
                                "IgstOnIntra": "N"
                            },
                            "DocDtls": {
                                "Typ": str(inv_type),
                                "No":  t["Invoice No"],
                                "Dt": str(doc_date)
                            },
                            "SellerDtls": {
                                "Gstin": "27AAFCD5862R013",#t["GSTIN of  Supplier"],
                                "LglNm": t["Supplier Name"],
                                "TrdNm": t["Supplier Name"],
                                "Addr1": t["Supplier Address"],
                                "Addr2": "",
                                "Loc": str(seller_loc), #'BIJNOR',#
                                "Pin": str(int(t["Supplier Pin Code"])), #'414204',#
                                "Stcd": t["GSTIN of  Supplier"][:2], #'27',#
                                "Ph": "",
                                "Em": ""
                            },
                            "BuyerDtls": {
                                "Gstin": t["GST Registration No"],
                                "LglNm": buyer["lgnm"],
                                "TrdNm": buyer["tradeNam"],
                                "Pos": t["GST Registration No"][:2],
                                "Addr1": add1,
                                "Addr2": add2,
                                "Loc": loc,
                                "Pin": pin,
                                "Stcd": t["GST Registration No"][:2],
                                "Ph": "",
                                "Em": ""
                            },
                            "DispDtls": {},
                            "ShipDtls": {},
                            "ValDtls": {
                                "AssVal": t["Item Price"],
                                "CgstVal": t["CGST Amount"],
                                "SgstVal": t["SGST Amount"],
                                "IgstVal": t["IGST Amount"],
                                "CesVal": "",
                                "StCesVal": "",
                                "Discount": 0,
                                "OthChrg": 0,
                                "RndOffAmt": 0,
                                "TotInvVal": t["Invoice Amount"],
                                "TotInvValFc": t["Invoice Amount"]    
                                }, 
                            "custom_fields": {},
                            "ItemList": [],
                            "ExpDtls": {},
                            "EwbDtls": {}              
                        }
                            # Initialize an empty list for items     
                    }
        except Exception as e: 
            ct.error_logger.error("Error while creating the json file:"+str(e))
            return buyer, "", ""
        invoice["transaction"]["ItemList"]=items
        # print(invoice)
        ct.info_logger.info(f"The json is: {invoice}")
        return invoice, customer_no, buyer_gstin 
    except Exception as e: 
        ct.error_logger.error("Error while creating the json file:"+str(e))
        # exit()
    

def email(doc_no, date, receiver_email):
    try:
        annexure_status = False
        email_status = False
        pdf_status = False
        # receiver_email = "hotel.accounts@cleartrip.com, rohan.d@cloudare.in"
        # date_obj = datetime.datetime.strptime(date, "%d/%m/%Y")
        # date = date_obj.strftime("%b")
        receiver_email = receiver_email.replace(";",",")
        print(receiver_email)
        # Convert the string to a datetime object
        # date_obj = datetime.datetime.strptime(date, "%d/%m/%Y")

        # Format the date to "Apr'23"
        # formatted_date = date_obj.strftime("%b'%y")
        formatted_date = date
        # Format the date to "Apr"
        

        # annexure(formatted_date1)
        # Email content
        sender_email = cg.sender_email
        subject = f"Commission Invoices for the month of {str(formatted_date)}"
        body = f"""Dear Team,

Please find the attached commission Invoices for the month of {str(formatted_date)}.

Regards,
{str(sender_email)}"""
        
        # Create the email
        message = MIMEMultipart()
        message["From"] = cg.sender_email
        message["To"] = receiver_email
        cc_emails = cg.cc_emails
        # message["Cc"] = ", ".join(cg.cc_emails)   
        message["Subject"] = subject

        # Attach the email body
        message.attach(MIMEText(body, "plain"))

        # Directory containing the PDF files
        

        # Generate a list of all PDF and XLSX files dynamically in the specified directory
        pdf_file_type = [str(doc_no) + ".pdf"]
        xlsx_file_type = [str(doc_no) + ".xlsx"]
        pdf_file_paths = []
        xlsx_file_paths = []
        for pdf_file in pdf_file_type:
            pdf_file_paths.extend(glob.glob(os.path.join(cg.pdf_directory, pdf_file)))
        for xlsx_file in xlsx_file_type:
            xlsx_file_paths.extend(glob.glob(os.path.join(cg.xlsx_directory, xlsx_file)))

        # Function to attach files dynamically
        if not pdf_file_paths:
            ct.info_logger.info(f"No PDF files found in '{cg.pdf_directory}'")
            # exit()
        # Function to attach files dynamically
        if not xlsx_file_paths:
            ct.info_logger.info(f"No PDF files found in '{cg.xlsx_directory}'")
            # exit()

        for file_path in pdf_file_paths:
            try:
                if any(pt in file_path for pt in pdf_file_type):
                    with open(file_path, "rb") as file:
                        # Create a MIMEBase object
                        attachment = MIMEBase("application", "octet-stream")
                        attachment.set_payload(file.read())

                    # Encode the file payload in Base64
                    encoders.encode_base64(attachment)

                    # Extract the filename dynamically from the file path
                    custom_file_name = os.path.basename(file_path)
                    ct.info_logger.info(f"The PDF file name is {custom_file_name}.")
                    # Add header to set custom file name
                    attachment.add_header(
                        "Content-Disposition",
                        f"attachment; filename={custom_file_name}",
                    )

                    # Attach the file to the message
                    message.attach(attachment)
                    pdf_status = True
            except FileNotFoundError:
                ct.error_logger.error(f"Error: File '{file_path}' not found to send the email.")
                pdf_status = False
                # exit()

        for file_path in xlsx_file_paths:
            try:
                if any(pt in file_path for pt in file_path):
                    with open(file_path, "rb") as file:
                        # Create a MIMEBase object
                        attachment = MIMEBase("application", "octet-stream")
                        attachment.set_payload(file.read())

                    # Encode the file payload in Base64
                    encoders.encode_base64(attachment)

                    # Extract the filename dynamically from the file path
                    custom_file_name = os.path.basename(file_path)
                    ct.info_logger.info(f"The annexure file name is {custom_file_name}.")
                    # Add header to set custom file name
                    attachment.add_header(
                        "Content-Disposition",
                        f"attachment; filename={custom_file_name}",
                    )

                    # Attach the file to the message
                    message.attach(attachment)
                    annexure_status = True
                else:
                    annexure_status = False
            except FileNotFoundError:
                ct.error_logger.error(f"Error: File '{file_path}' not found to send the email.")
                annexure_status = False
                # exit()

        # Combine all recipients (To + CC)
        # all_recipients = [receiver_email] + cc_emails

        all_recipients = []
        if receiver_email:  # Only add if not empty
            all_recipients.append(receiver_email)
        if cc_emails:
            # Filter out any empty strings in the CC list as well
            all_recipients.extend(email for email in cc_emails if email)

        try:
            email_status = False
            if not all_recipients:
                print("No valid recipients provided. Email will not be sent.")
            else:
            # Connect to SMTP server
                with smtplib.SMTP(cg.smtp_server, cg.port) as server:
                    server.starttls()  # Upgrade to secure connection
                    server.login(cg.sender_email, cg.password)  # Login to your email account
                    server.sendmail(sender_email, all_recipients, message.as_string())  # Send email
                    print("Email sent successfully!")
                    email_status = True
                ct.info_logger.info("The email has been sent successfully.")
            return annexure_status, email_status, pdf_status
        except Exception as e:
            ct.error_logger.error(f"Error occurred while sending the email: {e}")
            email_status = False
            return annexure_status, email_status, pdf_status
    except Exception as e:
            ct.error_logger.error(f"Error occurred while sending the email: {e}")
            email_status = False
            return annexure_status, email_status, pdf_status

def annexure():
    try:
        # Find all .xlsx files in the directory
        xlsx_files = glob.glob(os.path.join(cg.raw_annexure, "*.csv")) + glob.glob(os.path.join(cg.raw_annexure, "*.xlsx")) + glob.glob(os.path.join(cg.raw_annexure, "*.xlx"))
        print(xlsx_files)
        # Display the found files
        if xlsx_files:
            print("Found files:")
            for file in xlsx_files:
                print(file)
                file_path = file  # Replace with your file path

                # Read the Excel file into a DataFrame
                try:
                    if file.lower().endswith('.csv'):
                        df = pd.read_csv(file_path, encoding='latin1', on_bad_lines='skip')
                    elif file.lower().endswith('.xlsx') or file.lower().endswith('.xlx'):
                        df = pd.read_excel(file_path, engine='openpyxl')
                    else:
                        print(f"Unsupported file type: {file_path}. Skipping.")
                        continue
                except UnicodeDecodeError:
                    print(f"Error decoding file: {file_path}. Try a different encoding.")
                    continue
                print(df)
                print(df["Check In "])
                # Parse 'Booking Date' column
                if 'Booking Date' in df.columns:
                    try:
                        months = pd.to_datetime(df['Booking Date'], format="%d-%m-%Y", errors='coerce').dt.strftime("%b'%y")

                        # Format the date to "Apr'23"
                        # formatted_date = date_obj.strftime("%b'%y")
                #         df['Month'] = pd.to_datetime(df['Booking Date'], format="%d-%m-%Y", errors='coerce').dt.strftime("%b")
                    except Exception as e:
                        print(f"Error parsing 'Booking Date': {e}")
                        continue
                else:
                    print(f"'Booking Date' column not found in {file}. Skipping.")
                    continue
                # print(df['Month'])
                print(months)
                
                # Group data by 'Inv No' and process
                if 'Invoice No' in df.columns:
                    grouped = df.groupby('Invoice No')
                    for inv_no, group in grouped:
                        print(group)
                        customer_no = group['Customer No'].iloc[0] if 'Customer No' in group.columns else 'Unknown'
                        gst_no = group['GST No '].iloc[0] if 'GST No ' in group.columns else 'Unknown'
                        # month = group['Month'].iloc[0] if 'Month' in group.columns else 'Unknown'
                        month = months[0]
                        current_date = datetime.datetime.strptime(month, "%b'%y")

                        # Calculate the previous month
                        previous_month_date = current_date - datetime.timedelta(days=current_date.day)

                        # Format the previous month back to the desired format
                        previous_month = previous_month_date.strftime("%b'%y")
                        # print(month)
                        if 'Check In ' in group.columns:
                            # Convert the 'Check In' column to datetime if not already
                            group['Check In '] = pd.to_datetime(group['Check In '], errors='coerce').dt.strftime("%d-%m-%Y")
                            # print(group['Check In '])
                        if 'Check Out' in group.columns:
                            # Convert the 'Check In' column to datetime if not already
                            group['Check Out'] = pd.to_datetime(group['Check Out'], errors='coerce').dt.strftime("%d-%m-%Y")
                        if 'Trip ID ' in group.columns:
                            group['Trip ID '] = pd.to_numeric(group['Trip ID '], errors='coerce').fillna(0).astype(int)
                            print(group['Trip ID '])
                        
                        output_file = os.path.join(
                            cg.xlsx_directory, 
                            f"{str(customer_no)}_{str(gst_no)}_{str(previous_month)}.xlsx"
                        )
                        print(group)
                        print(f"Saving group to: {output_file}")
                        group.to_excel(output_file, index=False)

                        # Load the workbook to apply formatting
                        wb = load_workbook(output_file)
                        ws = wb["Sheet1"]

                        # Create a named style for number formatting
                        number_style = NamedStyle(name="number_style", number_format="0")

                        # Apply the style to the "Trip ID " column
                        trip_id_column = "J"  # Replace with the actual column if it's not B
                        for cell in ws[trip_id_column][1:]:  # Skip the header row
                            cell.style = number_style

                        # Save the workbook with formatting
                        wb.save(output_file)
                    # Move processed file to archive
                    print(str(file))
                    if os.path.exists(str(cg.archive_raw_annexure)+ str(file.replace("C:/codes/cleartrip/code/data/raw_annexure\\",""))):
                        os.remove(str(cg.archive_raw_annexure)+ str(file.replace("C:/codes/cleartrip/code/data/raw_annexure\\","")))
                    shutil.move(file, cg.archive_raw_annexure)
                    print(f"Moved file {file} to archive: {cg.archive_raw_annexure}")
                else:
                    print(f"'Invoice No' column not found in {file}. Skipping.")
        else:
            print("CSV not present.")
    except Exception as e:
        ct.error_logger.error(f"Error occurred while creating the annexure: {e}")
        
def delete_old_files():
    # Get the current time
    current_time = datetime.datetime.now()

    # Define the age limit (60 days)
    age_limit = timedelta(days=25)

    # Iterate over the files in the archive_raw_annexure directory
    for filename in os.listdir(cg.archive_raw_annexure):
        file_path = os.path.join(cg.archive_raw_annexure, filename)

        # Check if it's a file
        if os.path.isfile(file_path):
            # Get the last modified time of the file
            file_mtime = datetime.datetime.fromtimestamp(os.path.getmtime(file_path))

            # Check if the file is older than the age limit
            if current_time - file_mtime > age_limit:
                try:
                    # Delete the file
                    os.remove(file_path)
                    print(f"Deleted file: {file_path}")
                except Exception as e:
                    print(f"Error deleting file {file_path}: {e}")
    
    # Iterate over the files in the archive_raw_input directory
    for filename in os.listdir(cg.archive_raw_input):
        file_path = os.path.join(cg.archive_raw_input, filename)

        # Check if it's a file
        if os.path.isfile(file_path):
            # Get the last modified time of the file
            file_mtime = datetime.datetime.fromtimestamp(os.path.getmtime(file_path))

            # Check if the file is older than the age limit
            if current_time - file_mtime > age_limit:
                try:
                    # Delete the file
                    os.remove(file_path)
                    print(f"Deleted file: {file_path}")
                except Exception as e:
                    print(f"Error deleting file {file_path}: {e}")
    
    # Iterate over the files in the output directory
    for filename in os.listdir(cg.file):
        file_path = os.path.join(cg.file, filename)

        # Check if it's a file
        if os.path.isfile(file_path):
            # Get the last modified time of the file
            file_mtime = datetime.datetime.fromtimestamp(os.path.getmtime(file_path))

            # Check if the file is older than the age limit
            if current_time - file_mtime > age_limit:
                try:
                    # Delete the file
                    os.remove(file_path)
                    print(f"Deleted file: {file_path}")
                except Exception as e:
                    print(f"Error deleting file {file_path}: {e}")

     # Iterate over the files in the output directory
    for filename in os.listdir(cg.xlsx_directory):
        file_path = os.path.join(cg.xlsx_directory, filename)

        # Check if it's a file
        if os.path.isfile(file_path):
            # Get the last modified time of the file
            file_mtime = datetime.datetime.fromtimestamp(os.path.getmtime(file_path))

            # Check if the file is older than the age limit
            if current_time - file_mtime > age_limit:
                try:
                    # Delete the file
                    os.remove(file_path)
                    print(f"Deleted file: {file_path}")
                except Exception as e:
                    print(f"Error deleting file {file_path}: {e}")


def email_count(email_string):
    try:
        print(email_string)
        # Split the email IDs into a list
        email_list = email_string.split(',')

        # Create a new list to hold email IDs within the character limit
        adjusted_email_list = []
        current_length = 0
        limit = 100

        # Add email IDs while keeping the total length under the limit
        for email in email_list:
            if current_length + len(email) + (len(adjusted_email_list) > 0) <= limit:
                adjusted_email_list.append(email)
                current_length += len(email) + 1  # Add 1 for the comma

        # Join the adjusted list back into a string
        adjusted_email_string = ','.join(adjusted_email_list)
        print(adjusted_email_string)
        return adjusted_email_string
    except:
        return ""