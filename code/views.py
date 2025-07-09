
import datetime
import time
import os
import config as cg
import json_creator as jc
import pandas as pd
import cleartax as ct
import json
import shutil
import codecs
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle

invoice_data = {}

def mainProcess():
    # print("test")
    try:
        parts = []
        xlsx_date = ''
        for filename in os.listdir(cg.input_directory):
            if filename.endswith('.xlsx') or filename.endswith('.xlx'):
                ct.info_logger.info(f'Converting file {filename} to clear json.')
                df = pd.read_excel(str(cg.input_directory) + str(filename))
                # Add new columns with default or computed values
                df['DOC_STATUS'] = ''
                df['IRN'] = ''
                df['ACK_NO'] = ''
                df['ACK_DATE'] = ''
                df['SIGNED_INVOICE'] = ''
                df['QR_CODE'] = ''
                df['ERROR_MESSAGE'] = ''
                df['ANNEXURE_STATUS'] = ''
                df['EMAIL_STATUS'] = ''
                os.remove(str(cg.input_directory) + str(filename))
                # shutil.move((str(cg.input_directory) + str(filename)), cg.archive_raw_input)
                new_file = str(cg.output) + str(filename).replace(".xlsx","").replace(".xls","")+ str(".csv")
                i=1
                while (1==1):
                    try:
                        if os.path.isfile(new_file) and new_file.endswith('.csv'):
                            os.rename(new_file, str(cg.output) + str(filename).replace(".xlsx","").replace(".csv","") + "_" + str(i) + str(".csv"))
                        else:
                            ct.info_logger.info(f"New file has been crated in '{str(new_file)}'")
                            df.to_csv(new_file, index=False, encoding='utf-8')
                            print("Done")
                            break
                    except:
                        try:
                            i = i + 1
                        except:
                            pass
            elif filename.endswith('.csv'):
                ct.info_logger.info(f'Converting file {filename} to clear json.')
                df = pd.read_csv(str(cg.input_directory) + str(filename))

                # Add new columns with default or computed values
                df['DOC_STATUS'] = ''
                df['IRN'] = ''
                df['ACK_NO'] = ''
                df['ACK_DATE'] = ''
                df['SIGNED_INVOICE'] = ''
                df['QR_CODE'] = ''
                df['ERROR_MESSAGE'] = ''
                df['ANNEXURE_STATUS'] = ''
                df['EMAIL_STATUS'] = ''
                df['PDF_STATUS'] = ''
                os.remove(str(cg.input_directory) + str(filename))
                # shutil.move((str(cg.input_directory) + str(filename)), cg.archive_raw_input)
                new_file = str(cg.output) + str(filename).replace(".xlsx","").replace(".csv","")+ str(".csv")
                i=1
                while (1==1):
                    try:
                        if os.path.isfile(new_file) and new_file.endswith('.csv'):
                            os.rename(new_file, str(cg.output) + str(filename).replace(".xlsx","").replace(".csv","") + "_" + str(i) + str(".csv"))
                        else:
                            ct.info_logger.info(f"New file has been crated in '{str(new_file)}'")
                            df.to_csv(new_file, index=False, encoding='utf-8')
                            print("Done")
                            break
                    except:
                        try:
                            i = i + 1
                        except:
                            pass

        for filename in os.listdir(cg.output):
            if filename.endswith('.csv'): 
                ct.info_logger.info(f'Converting file {filename} to clear json.')
                df = pd.read_csv(str(cg.output) + str(filename))
                
                for i in range(0, len(df)):
                    try:
                        ct.info_logger.info(f"Invoice processing for {str(i)} out of {str(len(df))}.")
                        row = df.iloc[i]
                        t= row.to_dict()
                        invoice_number = t["Invoice No"]
                        print(invoice_number)
                        print(i)
                        date_obj = datetime.datetime.strptime(str(pd.Timestamp(t["Posting Date"]).strftime('%d/%m/%Y')), "%d/%m/%Y")

                        # Format the date to "Apr'23"
                        formatted_date = date_obj.strftime("%b'%y")
                        formatted_date = date_obj.strftime("%b'%y")
                        current_date = datetime.datetime.strptime(formatted_date, "%b'%y")

                        # Calculate the previous month
                        previous_month_date = current_date - datetime.timedelta(days=current_date.day)

                        # Format the previous month back to the desired format
                        previous_month = previous_month_date.strftime("%b'%y")
                        xlsx_date = previous_month.replace("'","_20")

                        invoice_data[f'{invoice_number}'], customer_no, buyer_gstin = jc.create_json(t)
                        print(invoice_data[f'{invoice_number}'])
                        payload = invoice_data[f'{invoice_number}']
                        irn, resp = ct.generate_IRN(payload)
                        # resp = json.load(resp)
                        print(t["Cust Email ID"])
                        reciever_email = str(t["Cust Email ID"]) if str(t["Cust Email ID"]) != 'nan' else ""
                        print(reciever_email)
                        print(resp)
                        inv_no = payload["transaction"]["DocDtls"]["No"]
                        
                        try:
                            if irn != "":
                                df.loc[df["Invoice No"] == str(inv_no), 'DOC_STATUS'] = resp[0]["document_status"]
                                df.loc[df["Invoice No"] == str(inv_no), 'IRN'] = irn
                                df.loc[df["Invoice No"] == str(inv_no), 'ACK_NO'] = str(int(resp[0]["govt_response"]["AckNo"]))
                                df.loc[df["Invoice No"] == str(inv_no), 'ACK_DATE'] = resp[0]["govt_response"]["AckDt"]
                                df.loc[df["Invoice No"] == str(inv_no), 'SIGNED_INVOICE'] = resp[0]["govt_response"]["SignedInvoice"]
                                df.loc[df["Invoice No"] == str(inv_no), 'QR_CODE'] = resp[0]["govt_response"]["SignedQRCode"]
                                print(resp)
                                # df.loc[df["Invoice No"] == str(inv_no), 'ERROR_MESSAGE'] = str(ct.error_message_list(resp))
                                response_pdf = ct.pdf(irn)
                                
                                print(response_pdf.status_code)
                                # Check if the response is successful (status code 200)
                                if response_pdf.status_code == 200:
                                    # Open a file in binary mode and write the content
                                    
                                    with open('data/pdf/'+str(customer_no) + "_" + str(buyer_gstin) + "_" + str(previous_month) +'.pdf', 'wb') as pdf_file:
                                        pdf_file.write(response_pdf.content)
                                    ct.info_logger.info(f"PDF saved successfully with status code is {response_pdf.status_code}.")
                                else:
                                    ct.error_logger(str(f"Failed to download PDF. Status code: {response_pdf.status_code}"))

                                annexure_status, email_status, pdf_status = jc.email(str(customer_no) + "_" + str(buyer_gstin) + "_" + str(previous_month),str(previous_month), reciever_email)
                                
                                if email_status == True:
                                    df.loc[df["Invoice No"] == str(inv_no), 'EMAIL_STATUS'] = "Success"
                                else:
                                    df.loc[df["Invoice No"] == str(inv_no), 'EMAIL_STATUS'] = "Failed"

                                if annexure_status == True:
                                    df.loc[df["Invoice No"] == str(inv_no), 'ANNEXURE_STATUS'] = "Success"
                                else:
                                    df.loc[df["Invoice No"] == str(inv_no), 'ANNEXURE_STATUS'] = "Failed"

                                if pdf_status == True:
                                    df.loc[df["Invoice No"] == str(inv_no), 'PDF_STATUS'] = "Success"
                                else:
                                    df.loc[df["Invoice No"] == str(inv_no), 'PDF_STATUS'] = "Failed"  
                            else:
                                df.loc[df["Invoice No"] == str(inv_no), 'DOC_STATUS'] = resp[0]["document_status"]
                                df.loc[df["Invoice No"] == str(inv_no), 'ERROR_MESSAGE'] = str(ct.error_message_list(resp))
                                df.loc[df["Invoice No"] == str(inv_no), 'EMAIL_STATUS'] = "Failed"
                                df.loc[df["Invoice No"] == str(inv_no), 'ANNEXURE_STATUS'] = "Failed"
                                df.loc[df["Invoice No"] == str(inv_no), 'PDF_STATUS'] = "Failed"
                        except Exception as e:
                            ct.info_logger.info(f"Error due to '{str(e)}'.")
                            df.loc[df["Invoice No"] == str(invoice_number), 'ERROR_MESSAGE'] = str(invoice_data[f'{invoice_number}'])
                            df.loc[df["Invoice No"] == str(invoice_number), 'DOC_STATUS'] = 'FAILED'
                    except Exception as e:
                            ct.info_logger.info(f"Error due to '{str(e)}'.")
                            df.loc[df["Invoice No"] == str(invoice_number), 'ERROR_MESSAGE'] = str(invoice_data[f'{invoice_number}'])
                            df.loc[df["Invoice No"] == str(invoice_number), 'DOC_STATUS'] = 'FAILED'
                            # df.loc[df["Invoice No"] == str(inv_no), 'ERROR_MESSAGE'] = str(invoice_data[f'{invoice_number}'])
                # shutil.move((str(cg.input_directory) + str("output/") + str(filename)), cg.archive_raw_input)
                # xlsx_date = previous_month.replace("'","_20")
                # shutil.move((str(cg.input_directory) + str(filename)), cg.archive_raw_input)
                ct.info_logger.info(f"The file has been removed from '{str(cg.output)}'")
                
                if '_' in filename:
                    base_name, extension = os.path.splitext(filename)
                    parts = base_name.split('_', 1)
                else:
                    parts = [filename]
                new_file = str(cg.final_output) + str(parts[0]).replace(".xlsx","").replace(".csv","") + "_" + str(xlsx_date) + str(".xlsx")
                i=1
                while (1==1):
                    try:
                        if os.path.isfile(new_file) and new_file.endswith('.xlsx'):
                            os.rename(new_file, str(cg.final_output) + str(filename).replace(".xlsx","_").replace(".csv","_") + str(xlsx_date) + "_" + str(i) + str(".xlsx"))
                        else:
                            ct.info_logger.info(f"New file has been crated in '{str(new_file)}'")
                            df.to_excel(new_file, index=False, float_format='%.0f')
                            os.remove(str(cg.output) + str(filename))
                            wb = load_workbook(new_file)
                            ws = wb["Sheet1"]

                            # Create a named style for number formatting
                            number_style = NamedStyle(name="number_style", number_format="0")

                            # Apply the style to the "Trip ID " column
                            trip_id_column = "F"  # Replace with the actual column if it's not B
                            for cell in ws[trip_id_column][1:]:  # Skip the header row
                                cell.style = number_style
                            trip_id_column = "H"  # Replace with the actual column if it's not B
                            for cell in ws[trip_id_column][1:]:  # Skip the header row
                                cell.style = number_style
                            trip_id_column = "I"  # Replace with the actual column if it's not B
                            for cell in ws[trip_id_column][1:]:  # Skip the header row
                                cell.style = number_style
                            trip_id_column = "AN"  # Replace with the actual column if it's not B
                            for cell in ws[trip_id_column][1:]:  # Skip the header row
                                cell.style = number_style

                            wb.save(new_file)
                            print("Done")
                            break
                    except Exception as e:
                        print(e)
                        try:
                            i = i + 1
                        except:
                            pass

            else:
                ct.error_logger.info(f'No excel file was found.')
    except Exception as e: 
        ct.error_logger.error(f"Exception:{str(e)}")
        if '_' in filename:
            base_name, extension = os.path.splitext(filename)
            parts = base_name.split('_', 1)
        else:
            parts = [filename]
        new_file = str(cg.final_output) + str(parts[0]).replace(".xlsx","_").replace(".csv","_") + str(xlsx_date) + str(".xlsx")
        i=1
        while (1==1):
            try:
                if os.path.isfile(new_file) and new_file.endswith('.xlsx'):
                    os.rename(new_file, str(cg.final_output) + str(filename).replace(".xlsx","_").replace(".csv","_") + "_" + str(i) + str(".xlsx"))
                else:
                    ct.info_logger.info(f"New file has been crated in '{str(new_file)}'")
                    df.to_excel(new_file, index=False, float_format='%.0f')
                    os.remove(str(cg.output) + str(filename))

                    wb = load_workbook(new_file)
                    ws = wb["Sheet1"]

                    # Create a named style for number formatting
                    number_style = NamedStyle(name="number_style", number_format="0")

                    # Apply the style to the "Trip ID " column
                    trip_id_column = "F"  # Replace with the actual column if it's not B
                    for cell in ws[trip_id_column][1:]:  # Skip the header row
                        cell.style = number_style
                    trip_id_column = "H"  # Replace with the actual column if it's not B
                    for cell in ws[trip_id_column][1:]:  # Skip the header row
                        cell.style = number_style
                    trip_id_column = "I"  # Replace with the actual column if it's not B
                    for cell in ws[trip_id_column][1:]:  # Skip the header row
                        cell.style = number_style
                    trip_id_column = "AN"  # Replace with the actual column if it's not B
                    for cell in ws[trip_id_column][1:]:  # Skip the header row
                        cell.style = number_style

                    wb.save(new_file)
                    print("Done")
                    break
            except:
                try:
                    i = i + 1
                except:
                    pass
        # shutil.move((str(cg.input_directory) + str(filename)), cg.archive_raw_input)
        
        print("Done")
        

